from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl import Workbook

def Mean(*x):
    return sum(x)/len(x)
def SD(*x):
    mean=Mean(*x)
    sd2=0
    for i in x:
        sd2=sd2+(i-mean)**2
    sd=(sd2/(len(x)-1))**0.5
    return sd
def SE(*x):
    mean=Mean(*x)
    sd=SD(*x)
    se=sd/(len(x)**0.5)
    return se
def take_data_safely(x,y,z):
    temp_list=[]
    for i in range(k):
        temp_list.append(x[y+i][z])
    temp_list=list(set(temp_list))
    if list(set(temp_list))==['']:
        if x[y]['Target Name']==内参基因:
            print('内参基因%s的%s没有有效数据'%(x[y]['Target Name'],z))
        else:
            print('基因%s的%s没有有效原始株数据'%(x[y]['Target Name'],z))
    else:
        if ('' in temp_list)==False:
            return temp_list[0]
        else:
            temp_list.remove('')
            return temp_list[0]
def check_too_few_repetitions(x,y,z):
    if z<=1:
        print('%s大组的基因%s有效样本过少，建议重做'%(x[y]['Sample Name'],x[y]['Target Name']))

print('要求：\n1.默认按行加样\n2.第一个大组必须是原始株\n3.大组内的第一个基因必须是内参基因\n4.每个大组内的基因顺序应相同\n5.你的原excel数据必须包括Well、Well Position、Omit、Sample Name、Target Name、Ct这几项')

#读取表格
path=filedialog.askopenfilename()
raw_data_document=load_workbook(path)
raw_data_sheet=raw_data_document.worksheets[0]
raw_data=list(raw_data_sheet.values)

#表格的每一行转化为一个词典，词典的键是列首
keys=raw_data[0]
data_dict_list=[]
for i in range(1,len(raw_data)):
    values=raw_data[i]
    data_dict=dict(list(zip(keys,values)))
    data_dict_list.append(data_dict)

#引入大组数、基因数、平行组数
m=eval(input('请输入实验大组数目：'))
原始株=input('请输入原始株的大组代号：')
n=eval(input('请输入内参基因在内的目标基因数目：'))
内参基因=input('请输入内参基因名称：')
k=eval(input('请输入平行加样数目：'))

#添加各组的Ct平均值和Ct样本标准差
keys=keys+('Ct Mean','Ct SD')
for a in range(m*n):
    小组=data_dict_list[(a*k):((a+1)*k)]
    accepted=[]
    for b in range(k):
        单孔=小组[b]
        if 单孔['Omit']==False:
            '''这组加样孔是有效的'''
            accepted.append(单孔['Ct'])
    check_too_few_repetitions(data_dict_list,a*k,len(accepted))
    Ct_Mean=Mean(*accepted)
    Ct_SD=SD(*accepted)
    for b in range(k):
        单孔=小组[b]
        if 单孔['Omit']==False:
            单孔.update({'Ct Mean':Ct_Mean,'Ct SD':Ct_SD})
        else:
            单孔.update({'Ct Mean':'','Ct SD':''})

#添加ΔCt、ΔCt平均值、ΔCt的样本标准差和ΔCt的样本标准误
keys=keys+('ΔCt','ΔCt Mean','ΔCt SD','ΔCt SE')
for a in range(m):
    大组=data_dict_list[(a*n*k):((a+1)*n*k)]
    for b in range(n):
        小组=大组[(b*k):((b+1)*k)]
        if 小组[0]['Target Name']==内参基因:
            '''这是内参基因组'''
            for c in range(k):
                单孔=小组[c]
                单孔.update({'ΔCt':'','ΔCt Mean':'','ΔCt SD':'','ΔCt SE':''})
        else:
            '''这不是内参基因'''
            accepted=[]
            #先加ΔCt
            for c in range(k):
                单孔=小组[c]
                if 单孔['Omit']==False:
                    '''这组加样孔是有效的'''
                    ΔCt=单孔['Ct']-take_data_safely(data_dict_list,a*n*k,'Ct Mean')
                    '''本空Ct值减去本大组内参的Ct平均值'''
                    accepted.append(ΔCt)
                else:
                    '''这组加样孔是无效的'''
                    ΔCt=''
                单孔.update({'ΔCt':ΔCt})
            #然后计算ΔCt平均值
            ΔCt_Mean=Mean(*accepted)
            #然后计算ΔCt的样本标准差、ΔCt的样本标准误，并加入ΔCt平均值、ΔCt的样本标准差、ΔCt的样本标准误
            for c in range(k):
                单孔=小组[c]
                if 单孔['Omit']==False:
                    '''这组加样孔是有效的'''
                    ΔCt_SD=(单孔['Ct SD']**2+data_dict_list[a*n*k]['Ct SD']**2)**0.5
                    '''两个独立数据相互作用后的结果的方差为两个独立数据的方差之和'''
                    ΔCt_SE=ΔCt_SD/(len(accepted)**0.5)
                    单孔.update({'ΔCt Mean':ΔCt_Mean,'ΔCt SD':ΔCt_SD,'ΔCt SE':ΔCt_SE})
                else:
                    '''这组加样孔是无效的'''
                    单孔.update({'ΔCt Mean':'','ΔCt SD':'','ΔCt SE':''})
                
#添加ΔΔCt、2^-ΔΔCt、2^ΔΔCt的平均值和2^ΔΔCt的样本标准差
keys=keys+('ΔΔCt','ΔΔCt Mean','2^(-ΔΔCt)','2^(-ΔΔCt) Mean','2^(-ΔΔCt) SD')
for a in range(m):
    大组=data_dict_list[(a*n*k):((a+1)*n*k)]
    if 大组[0]['Sample Name']==原始株:
        '''这是原始株'''
        for b in range(n*k):
            单孔=大组[b]
            单孔.update({'ΔΔCt':'','ΔΔCt Mean':'','2^(-ΔΔCt)':'','2^(-ΔΔCt) Mean':'','2^(-ΔΔCt) SD':''})
    else:
        '''这不是原始株'''
        for b in range(n):
            小组=大组[(b*k):((b+1)*k)]
            if 小组[0]['Target Name']==内参基因:
                '''这是内参基因组'''
                for c in range(k):
                    单孔=小组[c]
                    单孔.update({'ΔΔCt':'','ΔΔCt Mean':'','2^(-ΔΔCt)':'','2^(-ΔΔCt) Mean':'','2^(-ΔΔCt) SD':''})
            else:
                '''这不是内参基因'''
                #先加ΔΔCt
                accepted=[]
                for c in range(k):
                    单孔=小组[c]
                    if 单孔['Omit']==False:
                        '''这组加样孔是有效的'''
                        ΔΔCt=单孔['ΔCt']-take_data_safely(data_dict_list,b*k,'ΔCt Mean')
                        '''本组ΔCt值减去原始株对应基因的ΔCt的平均值'''
                        accepted.append(ΔΔCt)
                    else:
                        '''这组加样孔是无效的'''
                        ΔΔCt=''
                    单孔.update({'ΔΔCt':ΔΔCt})
                #然后计算并添加ΔΔCt的平均值
                ΔΔCt_Mean=Mean(*accepted)
                for c in range(k):
                    单孔=小组[c]
                    if 单孔['Omit']==False:
                        '''这组加样孔是有效的'''
                        单孔.update({'ΔΔCt Mean':ΔΔCt_Mean})
                    else:
                        '''这组加样孔是无效的'''
                        单孔.update({'ΔΔCt Mean':''})
                #然后加2^-ΔΔCt
                accepted=[]
                for c in range(k):
                    单孔=小组[c]
                    if 单孔['Omit']==False:
                        '''这组加样孔是有效的'''
                        指数化=2**(-单孔['ΔΔCt'])
                        accepted.append(指数化)
                    else:
                        '''这组加样孔是无效的'''
                        指数化=''
                    单孔.update({'2^(-ΔΔCt)':指数化})
                #然后计算并添加2^-ΔΔCt的平均值和2^-ΔΔCt的样本标准差
                指数化_Mean=Mean(*accepted)
                指数化_SD=SD(*accepted)
                for c in range(k):
                    单孔=小组[c]
                    if 单孔['Omit']==False:
                        '''这组加样孔是有效的'''
                        单孔.update({'2^(-ΔΔCt) Mean':指数化_Mean,'2^(-ΔΔCt) SD':指数化_SD})
                    else:
                        '''这组加样孔是无效的'''
                        单孔.update({'2^(-ΔΔCt) Mean':'','2^(-ΔΔCt) SD':''})

cooked_data=[keys]
for i in data_dict_list:
    cooked_data.append(tuple(i.values()))
cooked_data_document=Workbook()
cooked_data_sheet=cooked_data_document.worksheets[0]
for i in range(len(cooked_data)):
    for j in range(len(cooked_data[i])):
        cooked_data_sheet.cell(row=i+1,column=j+1,value=cooked_data[i][j])
cooked_data_document.save('C:/Users/ThinkPad/Desktop/test.xlsx')
print('文件已生成')
